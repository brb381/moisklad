from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from app.data_sources import MockSalesProvider
from app.dates import parse_month
from app.models import DailySales
from app.report import build_report
from app.stores import STORES, load_stores, resolve_store


def test_parse_month_boundaries():
    assert parse_month("2026-02") == (date(2026, 2, 1), date(2026, 2, 28))
    assert parse_month("2024-02") == (date(2024, 2, 1), date(2024, 2, 29))


def test_mock_provider_reads_existing_sample():
    start, end = parse_month("2026-05")
    data = MockSalesProvider().load_daily_sales("Планета", start, end)

    assert data[date(2026, 5, 1)].gross == 29834
    assert data[date(2026, 5, 1)].checks == 16
    assert len(data) == 31


def test_store_aliases_are_strict():
    assert resolve_store("planeta").name == "Планета"
    assert resolve_store("Июнь").code == "iyun"

    try:
        resolve_store("неизвестный магазин")
    except LookupError:
        pass
    else:
        raise AssertionError("unknown store must be rejected")


def test_stores_json_loads_expected_stores():
    stores = load_stores()
    assert [store.code for store in stores] == ["planeta", "iyun", "vysotnaya"]
    assert STORES[0].name == "Планета"


def test_build_report_writes_expected_cells():
    start, end = parse_month("2026-05")
    output = build_report(
        "Планета",
        start,
        end,
        {
            date(2026, 5, 1): DailySales(gross=12345.67, checks=12, positions=34),
            date(2026, 5, 2): DailySales(gross=5000.00, vat20=1000.00, checks=5, positions=8),
        },
        job_id="pytest",
    )
    wb = load_workbook(output, data_only=False)
    ws = wb.active

    assert ws["H13"].value == "01.05.2026"
    assert ws["I13"].value == "31.05.2026"
    assert ws["B20"].value == 12345.67
    assert ws["G21"].value == 4000
    assert ws["H20"].value == 12
    assert ws["I20"].value == 34
    assert ws["B51"].value == "=SUM(B20:B50)"
    assert ws.sheet_view.view == "normal"
    assert ws.sheet_view.selection[0].activeCell == "A1"


def test_mock_generated_report_matches_sample_table():
    start, end = parse_month("2026-05")
    data = MockSalesProvider().load_daily_sales("Планета", start, end)
    output = build_report("Планета", start, end, data, job_id="compare")

    sample = load_workbook("Форма_Отчета_о_валовом_обороте_2026_май_Планета.xlsx", data_only=False).active
    generated = load_workbook(output, data_only=False).active

    for row in range(20, 51):
        for col in (2, 3, 4, 5, 6, 7, 8, 9, 10):
            assert generated.cell(row, col).value == sample.cell(row, col).value

    for col in range(2, 11):
        assert generated.cell(51, col).value == sample.cell(51, col).value
