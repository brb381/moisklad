from __future__ import annotations

import calendar
from pathlib import Path

from openpyxl import load_workbook

from .config import ROOT_DIR, settings
from .models import DailySales
from .stores import find_template_for_store, resolve_store


MONTH_NAMES = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}


def choose_template(store: str) -> Path:
    return find_template_for_store(resolve_store(store))


def build_report(store: str, start, end, daily_sales: dict, job_id: str | None = None) -> Path:
    store_config = resolve_store(store)
    template = choose_template(store_config.name)
    wb = load_workbook(template)
    ws = wb.active

    ws["B3"] = settings.report_tenant
    ws["H3"] = settings.report_trade_name
    ws["B6"] = settings.report_rent_contract
    ws["H6"] = settings.report_room
    ws["B9"] = settings.report_tax_system
    ws["H9"] = settings.report_fiscal_operator
    ws["H13"] = start.strftime("%d.%m.%Y")
    ws["I13"] = end.strftime("%d.%m.%Y")

    last_day = calendar.monthrange(start.year, start.month)[1]
    for row in range(20, 51):
        day_num = row - 19
        if day_num <= last_day:
            day = start.replace(day=day_num)
            data: DailySales = daily_sales.get(day, DailySales())
            ws.cell(row, 1).value = day
            ws.cell(row, 2).value = data.gross
            ws.cell(row, 3).value = data.vat20
            ws.cell(row, 4).value = data.vat10
            ws.cell(row, 5).value = data.vat5
            ws.cell(row, 6).value = data.vat7
            ws.cell(row, 7).value = data.net
            ws.cell(row, 8).value = data.checks
            ws.cell(row, 9).value = data.positions
            ws.cell(row, 10).value = 0
        else:
            for col in range(1, 11):
                ws.cell(row, col).value = None

    total_row = 51
    for col in range(2, 11):
        letter = ws.cell(19, col).column_letter
        ws.cell(total_row, col).value = f"=SUM({letter}20:{letter}50)"

    month_name = MONTH_NAMES[start.month]
    safe_store = "".join(ch for ch in store_config.name if ch not in r'\/:*?"<>|').strip()
    output_dir = ROOT_DIR / "generated"
    output_dir.mkdir(exist_ok=True)
    suffix = f"_{job_id}" if job_id else ""
    output = output_dir / (
        f"Форма_Отчета_о_валовом_обороте_{start.year}_{month_name}_{safe_store}{suffix}.xlsx"
    )
    wb.save(output)
    return output
