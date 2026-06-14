from __future__ import annotations

import hashlib
import random
from abc import ABC, abstractmethod
from datetime import date

from openpyxl import load_workbook

from .config import ROOT_DIR, settings
from .dates import iter_days
from .models import DailySales
from .moysklad import MoySkladClient
from .stores import find_template_for_store, resolve_store


class SalesProvider(ABC):
    @abstractmethod
    def load_daily_sales(self, store: str, start: date, end: date) -> dict[date, DailySales]:
        raise NotImplementedError


class MoySkladSalesProvider(SalesProvider):
    def load_daily_sales(self, store: str, start: date, end: date) -> dict[date, DailySales]:
        return MoySkladClient().load_daily_sales(store, start, end)


class MockSalesProvider(SalesProvider):
    """Mock data source for responsible MVP tests before production API access.

    If a matching sample workbook exists, mock data is read from it. Otherwise a
    deterministic month dataset is generated from store + month.
    """

    def load_daily_sales(self, store: str, start: date, end: date) -> dict[date, DailySales]:
        store_config = resolve_store(store)
        sample = self._find_sample_workbook(store_config, start)
        if sample:
            return self._load_from_workbook(sample)
        return self._generate_deterministic(store_config.name, start, end)

    def _find_sample_workbook(self, store_config, start: date):
        month_lc = MONTH_NAMES[start.month].lower()
        path = find_template_for_store(store_config)
        if month_lc in path.stem.lower():
            return path
        return None

    def _load_from_workbook(self, path) -> dict[date, DailySales]:
        wb = load_workbook(path, data_only=False)
        ws = wb.active
        result: dict[date, DailySales] = {}
        for row in range(20, 51):
            day = ws.cell(row, 1).value
            if not day:
                continue
            if hasattr(day, "date"):
                day = day.date()
            result[day] = DailySales(
                gross=float(ws.cell(row, 2).value or 0),
                vat20=float(ws.cell(row, 3).value or 0),
                vat10=float(ws.cell(row, 4).value or 0),
                vat5=float(ws.cell(row, 5).value or 0),
                vat7=float(ws.cell(row, 6).value or 0),
                checks=int(ws.cell(row, 8).value or 0),
                positions=int(ws.cell(row, 9).value or 0),
            )
        return result

    def _generate_deterministic(self, store: str, start: date, end: date) -> dict[date, DailySales]:
        seed = int(hashlib.sha256(f"{store}:{start:%Y-%m}".encode()).hexdigest()[:12], 16)
        rng = random.Random(seed)
        result: dict[date, DailySales] = {}
        for day in iter_days(start, end):
            gross = round(rng.uniform(15000, 75000), 2)
            checks = rng.randint(15, 150)
            result[day] = DailySales(
                gross=gross,
                checks=checks,
                positions=checks,
            )
        return result


def get_sales_provider(source: str | None = None) -> SalesProvider:
    source = (source or settings.data_source).strip().lower()
    if source == "mock":
        return MockSalesProvider()
    if source == "moysklad":
        return MoySkladSalesProvider()
    raise ValueError("source must be 'mock' or 'moysklad'")


MONTH_NAMES = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}
